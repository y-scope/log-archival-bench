#!/usr/bin/env python3

import os
from pathlib import Path
import json
import numpy as np
import pandas as pd
import sys
from openpyxl.utils import get_column_letter

# Get the current working directory
current_dir = os.getcwd()

# Add it to sys.path
if current_dir not in sys.path:
    sys.path.append(current_dir)

from src.jsonsync import JsonItem


current_dir = Path(os.getcwd())

if os.path.basename(current_dir.resolve()) != "clp-bench-refactor":
    raise Exception("Must be run in clp-bench-refactor directory, if it was renamed, comment this line out")

assets_dir = current_dir / "assets"
outputs = [p for p in assets_dir.iterdir() if p.is_dir() if os.path.basename(str(p)) != "__pycache__"]

sheets = JsonItem({})

for output_dir in outputs:
    tool = os.path.basename(output_dir.resolve())
    filename = (output_dir / "output.json").resolve()
    if not os.path.exists(str(filename)):
        continue
    with open(filename, 'r') as file:
        output = json.load(file)
        for dataset, dataset_val in output.items():
            for configuration, configuration_val in dataset_val.items():

                methodology = f"{tool} ({configuration})"


                for key, value in configuration_val["ingest"].items():
                    if key == "memory_average_B":
                        sheets[f"{dataset} (ingest)"][methodology]["memory_average_MB"] = value / (1024*1024)
                    elif key == "compressed_size_B":
                        sheets[f"{dataset} (ingest)"][methodology]["compressed_size_MB"] = value / (1024*1024)
                    elif key == "decompressed_size_B":
                        sheets[f"{dataset} (ingest)"][methodology]["decompressed_size_MB"] = value / (1024*1024)
                    else:
                        sheets[f"{dataset} (ingest)"][methodology][key] = value
                try:
                    sheets[f"{dataset} (ingest)"][methodology]["compression_ratio"] = \
                            configuration_val["ingest"]["decompressed_size_B"] \
                            / configuration_val["ingest"]["compressed_size_B"]
                except ZeroDivisionError:
                    sheets[f"{dataset} (ingest)"][methodology]["compression_ratio"] = "compression failed: compressed size is 0"

                memory_cold = []
                memory_hot = []
                if "query_cold" in configuration_val and "query_hot" in configuration_val:
                    for key, value in enumerate(configuration_val["query_cold"]):
                        sheets[f"{dataset} (search)"][methodology][f"q{key}_cold"] = value["time_taken_s"]
                        sheets[f"{dataset} (result)"][methodology][f"q{key}_cold"] = value["result"]
                        if value["memory_average_B"] != -1:
                            memory_cold.append(value["memory_average_B"])

                    for key, value in enumerate(configuration_val["query_hot"]):
                        sheets[f"{dataset} (search)"][methodology][f"q{key}_hot"] = value["time_taken_s"]
                        sheets[f"{dataset} (result)"][methodology][f"q{key}_hot"] = value["result"]
                        if value["memory_average_B"] != -1:
                            memory_hot.append(value["memory_average_B"])

                if memory_cold:
                    memory_cold_val = sum(memory_cold)/len(memory_cold) / (1024*1024)
                    sheets[f"{dataset} (search)"][methodology]["memory_average_cold_MB"] = memory_cold_val

                if memory_hot:
                    memory_hot_val = (sum(memory_hot)/len(memory_hot)) / (1024*1024)
                    sheets[f"{dataset} (search)"][methodology]["memory_average_hot_MB"] = memory_hot_val


with pd.ExcelWriter("clp_bench.xlsx") as writer:
    for sheet_name, data in sheets.items():
        df = pd.DataFrame(data.compile())
        df.sort_index()

        df.to_excel(writer, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        for i, column in enumerate(df.columns):
            column_width = max(df[column].astype(str).map(len).max(), len(column))
            worksheet.column_dimensions[get_column_letter(i+2)].width = column_width + 6  # +1 for 1 indexing, +1 more for column that titles rows
        worksheet.column_dimensions[get_column_letter(1)].width = 30
